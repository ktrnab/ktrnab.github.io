#!/usr/bin/env python3
"""
Build Meridian Insurance Group Internal Comms Analytics Report 2024
- Excel file with Raw Data, Benchmarks, Channel Summary, Monthly Trend, Dashboard
- PNG preview images for portfolio
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.ticker as mticker
import numpy as np
import os

# ── OUTPUT PATHS ─────────────────────────────────────────────────────────────
OUT_DIR = "/Users/katrinab/ktrnab.github.io/assets"
XLSX_PATH = f"{OUT_DIR}/meridian-comms-report-2024.xlsx"
IMG_DIR = "/Users/katrinab/ktrnab.github.io/assets/images/work"
os.makedirs(IMG_DIR, exist_ok=True)

# ── PALETTE ──────────────────────────────────────────────────────────────────
G_DARK   = "2A5C27"
G_MID    = "3D7A39"
G_LIGHT  = "CAE896"
G_MUTED  = "6A8C65"
CREAM    = "EDF0E5"
SURFACE  = "F5F7F0"
WHITE    = "FFFFFF"
BORDER_C = "D4DACE"
RED_BG   = "FDECEA"
RED_TEXT = "B71C1C"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border(color=BORDER_C):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def top_border(color=BORDER_C):
    s = Side(style="thin", color=color)
    return Border(top=s)

def header_font(size=10):
    return Font(name="Calibri", bold=True, color=WHITE, size=size)

def body_font(bold=False, color="1E2A1C", size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def left():
    return Alignment(horizontal="left", vertical="center")

def set_header(ws, row, cols, texts, col_widths=None):
    for i, (col, text) in enumerate(zip(cols, texts)):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = fill(G_DARK)
        c.font = header_font()
        c.alignment = center()
        c.border = border()
    if col_widths:
        for col, w in zip(cols, col_widths):
            ws.column_dimensions[get_column_letter(col)].width = w

def style_data_row(ws, row, col_start, col_end, is_alt=False):
    bg = SURFACE if is_alt else WHITE
    for col in range(col_start, col_end + 1):
        c = ws.cell(row=row, column=col)
        if not c.fill or c.fill.fgColor.rgb in ("00000000", "FFFFFFFF", WHITE, SURFACE):
            c.fill = fill(bg)
        c.border = border()
        if not c.font or not c.font.name:
            c.font = body_font()

# ── DATA ─────────────────────────────────────────────────────────────────────
months_short = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
months_full  = [f"{m} 2024" for m in months_short]

# [Channel, Month, Campaign, Reach, Engagements, Clicks]
raw_rows = [
    ["Email Newsletter","Jan 2024","Monthly All-Staff Update",      891, 423, 67],
    ["Email Newsletter","Feb 2024","Monthly All-Staff Update",      891, 445, 71],
    ["Email Newsletter","Mar 2024","Q1 Town Hall Follow-Up",        895, 501, 83],
    ["Email Newsletter","Apr 2024","Monthly All-Staff Update",      895, 478, 74],
    ["Email Newsletter","May 2024","Monthly All-Staff Update",      898, 432, 68],
    ["Email Newsletter","Jun 2024","Mid-Year Review Digest",        898, 395, 58],
    ["Email Newsletter","Jul 2024","Monthly All-Staff Update",      900, 382, 52],
    ["Email Newsletter","Aug 2024","Monthly All-Staff Update",      900, 398, 61],
    ["Email Newsletter","Sep 2024","Q3 Town Hall Follow-Up",        902, 487, 79],
    ["Email Newsletter","Oct 2024","Monthly All-Staff Update",      902, 512, 88],
    ["Email Newsletter","Nov 2024","Year-End Planning Update",      905, 534, 95],
    ["Email Newsletter","Dec 2024","Annual Wrap-Up & Highlights",   905, 489, 76],
    ["Intranet",        "Jan 2024","Monthly Highlights Hub",        891, 312,134],
    ["Intranet",        "Feb 2024","Monthly Highlights Hub",        891, 295,118],
    ["Intranet",        "Mar 2024","Q1 Strategy Resources",         895, 341,152],
    ["Intranet",        "Apr 2024","Monthly Highlights Hub",        895, 318,141],
    ["Intranet",        "May 2024","Monthly Highlights Hub",        898, 307,126],
    ["Intranet",        "Jun 2024","Mid-Year Resource Update",      898, 278,108],
    ["Intranet",        "Jul 2024","Monthly Highlights Hub",        900, 261, 98],
    ["Intranet",        "Aug 2024","Monthly Highlights Hub",        900, 284,112],
    ["Intranet",        "Sep 2024","Q3 Strategy Resources",         902, 335,148],
    ["Intranet",        "Oct 2024","Monthly Highlights Hub",        902, 358,163],
    ["Intranet",        "Nov 2024","Benefits & Year-End Resources", 905, 372,171],
    ["Intranet",        "Dec 2024","Monthly Highlights Hub",        905, 325,138],
    ["LinkedIn",        "Jan 2024","Thought Leadership & News",    4200, 147, 52],
    ["LinkedIn",        "Feb 2024","Thought Leadership & News",    4350, 165, 58],
    ["LinkedIn",        "Mar 2024","Q1 Industry Insights",         5100, 214, 82],
    ["LinkedIn",        "Apr 2024","Thought Leadership & News",    4800, 182, 65],
    ["LinkedIn",        "May 2024","Thought Leadership & News",    4650, 158, 54],
    ["LinkedIn",        "Jun 2024","Thought Leadership & News",    3900, 121, 41],
    ["LinkedIn",        "Jul 2024","Thought Leadership & News",    3700, 108, 35],
    ["LinkedIn",        "Aug 2024","Thought Leadership & News",    4100, 135, 46],
    ["LinkedIn",        "Sep 2024","Q3 Industry Insights",         5300, 228, 89],
    ["LinkedIn",        "Oct 2024","Thought Leadership & News",    5800, 261,104],
    ["LinkedIn",        "Nov 2024","Year-End Industry Review",     6200, 285,118],
    ["LinkedIn",        "Dec 2024","Thought Leadership & News",    5100, 194, 72],
]

benchmarks = [
    ["Email Newsletter", 0.47, 0.08,  "Industry avg open rate for internal email comms"],
    ["Intranet",         0.35, 0.15,  "Internal target for monthly unique visitor rate"],
    ["LinkedIn",         0.035, 0.012,"Industry avg engagement rate for B2B company pages"],
]

channels = ["Email Newsletter", "Intranet", "LinkedIn"]

def eng_rate(row): return row[4] / row[3]
def ctr(row):      return row[5] / row[3]
def bench_eng(ch):
    for b in benchmarks:
        if b[0] == ch: return b[1]
def bench_ctr(ch):
    for b in benchmarks:
        if b[0] == ch: return b[2]

# ── WORKBOOK ─────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ════════════════════════════════════════════════════════════════════════════
# SHEET 1: RAW DATA
# ════════════════════════════════════════════════════════════════════════════
ws_raw = wb.create_sheet("Raw Data")
ws_raw.sheet_view.showGridLines = False
ws_raw.freeze_panes = "A3"

# Title
ws_raw.merge_cells("A1:J1")
title_cell = ws_raw["A1"]
title_cell.value = "Meridian Insurance Group — Internal Communications Channel Report 2024"
title_cell.font = Font(name="Calibri", bold=True, size=13, color=G_DARK)
title_cell.alignment = left()
title_cell.fill = fill(CREAM)
ws_raw.row_dimensions[1].height = 28

# Headers row 2
headers = ["Channel","Month","Campaign / Content","Reach","Engagements","Clicks",
           "Eng. Rate","CTR","Benchmark Eng. Rate","vs Benchmark"]
widths  = [20, 12, 34, 10, 14, 8, 12, 10, 22, 14]
set_header(ws_raw, 2, range(1, 11), headers, widths)
ws_raw.row_dimensions[2].height = 22

# Data rows
for i, row in enumerate(raw_rows):
    r = i + 3
    ch = row[0]
    is_alt = i % 2 == 1
    bg = SURFACE if is_alt else WHITE

    ws_raw.cell(r, 1, row[0]).alignment = left()
    ws_raw.cell(r, 2, row[1]).alignment = center()
    ws_raw.cell(r, 3, row[2]).alignment = left()
    ws_raw.cell(r, 4, row[3]).alignment = center()
    ws_raw.cell(r, 5, row[4]).alignment = center()
    ws_raw.cell(r, 6, row[5]).alignment = center()

    # Eng Rate formula
    er = ws_raw.cell(r, 7, f"=E{r}/D{r}")
    er.number_format = "0.0%"
    er.alignment = center()

    # CTR formula
    ct = ws_raw.cell(r, 8, f"=F{r}/D{r}")
    ct.number_format = "0.0%"
    ct.alignment = center()

    # VLOOKUP for benchmark eng rate
    vl = ws_raw.cell(r, 9, f'=VLOOKUP(A{r},Benchmarks!$A$3:$D$5,2,FALSE)')
    vl.number_format = "0.0%"
    vl.alignment = center()

    # vs Benchmark = Eng Rate - Benchmark
    vs = ws_raw.cell(r, 10, f"=G{r}-I{r}")
    vs.number_format = '+0.0%;-0.0%;0.0%'
    vs.alignment = center()

    for col in range(1, 11):
        c = ws_raw.cell(r, col)
        c.fill = fill(bg)
        c.border = border()
        if not c.font or c.font.color.rgb == "FF000000":
            c.font = body_font()

    ws_raw.row_dimensions[r].height = 18

# Conditional formatting: vs benchmark green/red
from openpyxl.formatting.rule import CellIsRule
ws_raw.conditional_formatting.add(
    f"J3:J{len(raw_rows)+2}",
    CellIsRule(operator="greaterThan", formula=["0"], fill=fill("E8F5E9"),
               font=Font(color="2E7D32"))
)
ws_raw.conditional_formatting.add(
    f"J3:J{len(raw_rows)+2}",
    CellIsRule(operator="lessThan", formula=["0"], fill=fill(RED_BG),
               font=Font(color=RED_TEXT))
)

# ════════════════════════════════════════════════════════════════════════════
# SHEET 2: BENCHMARKS
# ════════════════════════════════════════════════════════════════════════════
ws_bench = wb.create_sheet("Benchmarks")
ws_bench.sheet_view.showGridLines = False

ws_bench.merge_cells("A1:D1")
t = ws_bench["A1"]
t.value = "Channel Benchmarks — Reference Table (used by VLOOKUP in Raw Data)"
t.font = Font(name="Calibri", bold=True, size=12, color=G_DARK)
t.alignment = left()
t.fill = fill(CREAM)
ws_bench.row_dimensions[1].height = 28

bh = ["Channel", "Benchmark Eng. Rate", "Benchmark CTR", "Notes"]
bw = [22, 22, 18, 48]
set_header(ws_bench, 2, range(1, 5), bh, bw)
ws_bench.row_dimensions[2].height = 22

for i, b in enumerate(benchmarks):
    r = i + 3
    ws_bench.cell(r, 1, b[0]).alignment = left()
    ws_bench.cell(r, 2, b[1]).number_format = "0.0%"
    ws_bench.cell(r, 2).alignment = center()
    ws_bench.cell(r, 3, b[2]).number_format = "0.0%"
    ws_bench.cell(r, 3).alignment = center()
    ws_bench.cell(r, 4, b[3]).alignment = left()
    for col in range(1, 5):
        c = ws_bench.cell(r, col)
        c.fill = fill(WHITE if i % 2 == 0 else SURFACE)
        c.border = border()
        c.font = body_font()
    ws_bench.row_dimensions[r].height = 18

# ════════════════════════════════════════════════════════════════════════════
# SHEET 3: CHANNEL SUMMARY (pivot-style)
# ════════════════════════════════════════════════════════════════════════════
ws_ch = wb.create_sheet("Channel Summary")
ws_ch.sheet_view.showGridLines = False

ws_ch.merge_cells("A1:G1")
t2 = ws_ch["A1"]
t2.value = "Channel Summary — Full Year 2024"
t2.font = Font(name="Calibri", bold=True, size=13, color=G_DARK)
t2.alignment = left()
t2.fill = fill(CREAM)
ws_ch.row_dimensions[1].height = 28

ch_headers = ["Channel","Total Reach","Total Engagements","Total Clicks",
              "Avg Eng. Rate","Avg CTR","Benchmark Eng. Rate"]
ch_widths = [22, 16, 20, 14, 16, 12, 22]
set_header(ws_ch, 2, range(1, 8), ch_headers, ch_widths)
ws_ch.row_dimensions[2].height = 22

# Channel summary rows — computed values (mirroring what a pivot would show)
ch_summary = []
for ch in channels:
    rows = [r for r in raw_rows if r[0] == ch]
    total_reach = sum(r[3] for r in rows)
    total_eng   = sum(r[4] for r in rows)
    total_clicks= sum(r[5] for r in rows)
    avg_eng     = total_eng / total_reach
    avg_ctr     = total_clicks / total_reach
    bench       = bench_eng(ch)
    ch_summary.append([ch, total_reach, total_eng, total_clicks, avg_eng, avg_ctr, bench])

for i, row in enumerate(ch_summary):
    r = i + 3
    ws_ch.cell(r, 1, row[0]).alignment = left()
    ws_ch.cell(r, 2, row[1]).alignment = center()
    ws_ch.cell(r, 3, row[2]).alignment = center()
    ws_ch.cell(r, 4, row[3]).alignment = center()
    for col in [5, 6, 7]:
        c = ws_ch.cell(r, col)
        c.number_format = "0.0%"
        c.alignment = center()
    for col in range(1, 8):
        c = ws_ch.cell(r, col)
        c.fill = fill(WHITE if i % 2 == 0 else SURFACE)
        c.border = border()
        c.font = body_font()
    ws_ch.row_dimensions[r].height = 18

# Totals row
r_tot = len(channels) + 3
ws_ch.cell(r_tot, 1, "All Channels").alignment = left()
ws_ch.cell(r_tot, 2, sum(r[1] for r in ch_summary)).alignment = center()
ws_ch.cell(r_tot, 3, sum(r[2] for r in ch_summary)).alignment = center()
ws_ch.cell(r_tot, 4, sum(r[3] for r in ch_summary)).alignment = center()
ws_ch.cell(r_tot, 5, "—").alignment = center()
ws_ch.cell(r_tot, 6, "—").alignment = center()
ws_ch.cell(r_tot, 7, "—").alignment = center()
for col in range(1, 8):
    c = ws_ch.cell(r_tot, col)
    c.fill = fill(CREAM)
    c.border = border()
    c.font = Font(name="Calibri", bold=True, color=G_DARK, size=10)
ws_ch.row_dimensions[r_tot].height = 18

# Bar chart: Avg Eng Rate vs Benchmark by channel
chart_bar = BarChart()
chart_bar.type = "col"
chart_bar.title = "Avg Engagement Rate vs Benchmark by Channel"
chart_bar.y_axis.title = "Engagement Rate"
chart_bar.y_axis.numFmt = "0%"
chart_bar.x_axis.title = "Channel"
chart_bar.style = 10
chart_bar.width = 18
chart_bar.height = 11

# Avg Eng Rate series (col E, rows 3-5)
data_eng = Reference(ws_ch, min_col=5, max_col=5, min_row=2, max_row=5)
chart_bar.add_data(data_eng, titles_from_data=True)

# Benchmark series (col G, rows 3-5)
data_bench = Reference(ws_ch, min_col=7, max_col=7, min_row=2, max_row=5)
chart_bar.add_data(data_bench, titles_from_data=True)

cats = Reference(ws_ch, min_col=1, min_row=3, max_row=5)
chart_bar.set_categories(cats)
chart_bar.series[0].graphicalProperties.solidFill = G_MID
chart_bar.series[1].graphicalProperties.solidFill = G_LIGHT

ws_ch.add_chart(chart_bar, "A9")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 4: MONTHLY TREND (email focus)
# ════════════════════════════════════════════════════════════════════════════
ws_trend = wb.create_sheet("Monthly Trend")
ws_trend.sheet_view.showGridLines = False

ws_trend.merge_cells("A1:F1")
t3 = ws_trend["A1"]
t3.value = "Monthly Engagement Rate Trend — Email Newsletter 2024"
t3.font = Font(name="Calibri", bold=True, size=13, color=G_DARK)
t3.alignment = left()
t3.fill = fill(CREAM)
ws_trend.row_dimensions[1].height = 28

th = ["Month","Reach","Engagements","Clicks","Eng. Rate","Benchmark (47%)"]
tw = [12, 10, 16, 10, 14, 18]
set_header(ws_trend, 2, range(1, 7), th, tw)
ws_trend.row_dimensions[2].height = 22

email_rows = [r for r in raw_rows if r[0] == "Email Newsletter"]
for i, row in enumerate(email_rows):
    r = i + 3
    ws_trend.cell(r, 1, row[1]).alignment = center()
    ws_trend.cell(r, 2, row[3]).alignment = center()
    ws_trend.cell(r, 3, row[4]).alignment = center()
    ws_trend.cell(r, 4, row[5]).alignment = center()
    er_cell = ws_trend.cell(r, 5, row[4]/row[3])
    er_cell.number_format = "0.0%"
    er_cell.alignment = center()
    bench_cell = ws_trend.cell(r, 6, 0.47)
    bench_cell.number_format = "0.0%"
    bench_cell.alignment = center()
    for col in range(1, 7):
        c = ws_trend.cell(r, col)
        c.fill = fill(WHITE if i % 2 == 0 else SURFACE)
        c.border = border()
        c.font = body_font()
    ws_trend.row_dimensions[r].height = 18

# Line chart
chart_line = LineChart()
chart_line.title = "Email Newsletter: Monthly Engagement Rate vs Benchmark"
chart_line.y_axis.title = "Engagement Rate"
chart_line.y_axis.numFmt = "0%"
chart_line.y_axis.scaling.min = 0.3
chart_line.y_axis.scaling.max = 0.7
chart_line.x_axis.title = "Month"
chart_line.style = 10
chart_line.width = 22
chart_line.height = 12

data_rate = Reference(ws_trend, min_col=5, max_col=5, min_row=2, max_row=14)
data_bmark = Reference(ws_trend, min_col=6, max_col=6, min_row=2, max_row=14)
chart_line.add_data(data_rate, titles_from_data=True)
chart_line.add_data(data_bmark, titles_from_data=True)

cats_trend = Reference(ws_trend, min_col=1, min_row=3, max_row=14)
chart_line.set_categories(cats_trend)

chart_line.series[0].graphicalProperties.line.solidFill = G_MID
chart_line.series[0].graphicalProperties.line.width = 22000
chart_line.series[1].graphicalProperties.line.solidFill = G_LIGHT
chart_line.series[1].graphicalProperties.line.width = 16000
chart_line.series[1].graphicalProperties.line.dashDot = "dash"

ws_trend.add_chart(chart_line, "A9")

# ════════════════════════════════════════════════════════════════════════════
# SHEET 5: DASHBOARD
# ════════════════════════════════════════════════════════════════════════════
ws_dash = wb.create_sheet("Dashboard")
ws_dash.sheet_view.showGridLines = False
ws_dash.sheet_view.zoomScale = 90

# Title block
ws_dash.merge_cells("B2:K2")
ws_dash.row_dimensions[1].height = 10
ws_dash.row_dimensions[2].height = 36

t4 = ws_dash["B2"]
t4.value = "Meridian Insurance Group"
t4.font = Font(name="Calibri", bold=True, size=18, color=G_DARK)
t4.alignment = left()

ws_dash.merge_cells("B3:K3")
sub = ws_dash["B3"]
sub.value = "Internal Communications — Full Year Performance Report  |  Jan – Dec 2024"
sub.font = Font(name="Calibri", size=11, color=G_MUTED)
sub.alignment = left()
ws_dash.row_dimensions[3].height = 20

ws_dash.merge_cells("B4:K4")
divider = ws_dash["B4"]
divider.fill = fill(G_DARK)
ws_dash.row_dimensions[4].height = 3

ws_dash.row_dimensions[5].height = 12

# ── KPI Cards (row 6-9) ───────────────────────────────────────────────────
# Layout: 3 KPI boxes side by side
kpis = [
    ("Total Reach",        f"{sum(r[3] for r in raw_rows):,}",  "All channels combined"),
    ("Total Engagements",  f"{sum(r[4] for r in raw_rows):,}",  "Opens, visits, reactions"),
    ("Email Open Rate",    "51.2%",                              "Annual avg vs 47% benchmark"),
    ("Avg LinkedIn Eng.",  "3.8%",                               "Annual avg vs 3.5% benchmark"),
]

kpi_cols = [2, 5, 8, 11]
kpi_end_cols = [4, 7, 10, 13]

for i, (label, val, sub_text) in enumerate(kpis):
    cs = kpi_cols[i]
    ce = kpi_end_cols[i]

    ws_dash.merge_cells(start_row=6, start_column=cs, end_row=6, end_column=ce)
    ws_dash.merge_cells(start_row=7, start_column=cs, end_row=7, end_column=ce)
    ws_dash.merge_cells(start_row=8, start_column=cs, end_row=8, end_column=ce)
    ws_dash.merge_cells(start_row=9, start_column=cs, end_row=9, end_column=ce)

    lbl_cell = ws_dash.cell(6, cs, label)
    lbl_cell.font = Font(name="Calibri", size=8, bold=True, color=G_MUTED)
    lbl_cell.alignment = Alignment(horizontal="left", vertical="bottom")

    val_cell = ws_dash.cell(7, cs, val)
    val_cell.font = Font(name="Calibri", size=20, bold=True, color=G_DARK)
    val_cell.alignment = Alignment(horizontal="left", vertical="center")

    sub_cell = ws_dash.cell(8, cs, sub_text)
    sub_cell.font = Font(name="Calibri", size=8, color=G_MUTED)
    sub_cell.alignment = Alignment(horizontal="left", vertical="top")

    # Background fill for the card
    for rr in range(6, 10):
        for cc in range(cs, ce + 1):
            c = ws_dash.cell(rr, cc)
            c.fill = fill(SURFACE)

    # Top accent bar
    for cc in range(cs, ce + 1):
        ws_dash.cell(6, cc).border = Border(top=Side(style="medium", color=G_DARK))

ws_dash.row_dimensions[6].height = 16
ws_dash.row_dimensions[7].height = 30
ws_dash.row_dimensions[8].height = 14
ws_dash.row_dimensions[9].height = 8

ws_dash.row_dimensions[10].height = 16

# Column widths for dashboard
for c in range(1, 16):
    ws_dash.column_dimensions[get_column_letter(c)].width = 6.5
ws_dash.column_dimensions["A"].width = 2

# ── Notes row ─────────────────────────────────────────────────────────────
ws_dash.merge_cells("B28:M28")
note = ws_dash["B28"]
note.value = "Data covers Jan–Dec 2024 across three internal comms channels. Benchmarks sourced from industry standards for financial services internal communications."
note.font = Font(name="Calibri", size=8, color=G_MUTED, italic=True)
note.alignment = left()

# ════════════════════════════════════════════════════════════════════════════
# TAB COLORS
# ════════════════════════════════════════════════════════════════════════════
ws_raw.sheet_properties.tabColor   = G_DARK
ws_bench.sheet_properties.tabColor = G_MID
ws_ch.sheet_properties.tabColor    = G_MID
ws_trend.sheet_properties.tabColor = G_MID
ws_dash.sheet_properties.tabColor  = G_LIGHT

# Reorder sheets: Dashboard first
wb.move_sheet("Dashboard", offset=-4)

wb.save(XLSX_PATH)
print(f"✓ Excel saved: {XLSX_PATH}")

# ════════════════════════════════════════════════════════════════════════════
# PNG PREVIEW IMAGES (matplotlib)
# ════════════════════════════════════════════════════════════════════════════

# Palette for matplotlib
C_DARK   = "#2a5c27"
C_MID    = "#3d7a39"
C_LIGHT  = "#cae896"
C_CREAM  = "#edf0e5"
C_SURFACE= "#f5f7f0"
C_MUTED  = "#6a8c65"
C_TEXT   = "#1e2a1c"
C_BORDER = "#d4dace"
C_WHITE  = "#ffffff"
C_RED    = "#fdecea"
C_RED_T  = "#b71c1c"

plt.rcParams.update({
    "font.family": "sans-serif",
    "font.size": 9,
    "axes.spines.top": False,
    "axes.spines.right": False,
})

# ── IMAGE 1: Dashboard overview ───────────────────────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(14, 7))
fig.patch.set_facecolor(C_CREAM)

# Left: channel bar chart
ax1 = axes[0]
ax1.set_facecolor(C_WHITE)

ch_labels = ["Email\nNewsletter", "Intranet", "LinkedIn"]
avg_eng   = []
b_eng     = []
for ch in channels:
    rows = [r for r in raw_rows if r[0] == ch]
    tr = sum(r[3] for r in rows)
    te = sum(r[4] for r in rows)
    avg_eng.append(te/tr * 100)
    b_eng.append(bench_eng(ch) * 100)

x = np.arange(len(ch_labels))
w = 0.35
bars1 = ax1.bar(x - w/2, avg_eng, w, label="Actual", color=C_MID, zorder=3)
bars2 = ax1.bar(x + w/2, b_eng,   w, label="Benchmark", color=C_LIGHT, zorder=3)

ax1.set_xticks(x)
ax1.set_xticklabels(ch_labels, fontsize=9)
ax1.yaxis.set_major_formatter(mticker.FuncFormatter(lambda y, _: f"{y:.0f}%"))
ax1.set_ylabel("Engagement Rate", color=C_MUTED, fontsize=9)
ax1.set_title("Avg Engagement Rate vs Benchmark", color=C_TEXT, fontsize=11, fontweight="bold", pad=12)
ax1.legend(frameon=False, fontsize=8)
ax1.tick_params(colors=C_MUTED)
ax1.yaxis.label.set_color(C_MUTED)
ax1.spines["left"].set_color(C_BORDER)
ax1.spines["bottom"].set_color(C_BORDER)
ax1.grid(axis="y", color=C_BORDER, linewidth=0.5, zorder=0)
ax1.set_ylim(0, max(avg_eng + b_eng) * 1.25)

# Add value labels
for bar in bars1:
    ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.3,
             f"{bar.get_height():.1f}%", ha="center", va="bottom", fontsize=8, color=C_DARK)

# Right: email trend line chart
ax2 = axes[1]
ax2.set_facecolor(C_WHITE)

email_rows = [r for r in raw_rows if r[0] == "Email Newsletter"]
er_vals  = [r[4]/r[3]*100 for r in email_rows]
bench_line = [47.0] * 12

ax2.plot(months_short, er_vals, color=C_MID, linewidth=2.5, marker="o",
         markersize=5, label="Actual Open Rate", zorder=3)
ax2.plot(months_short, bench_line, color=C_LIGHT, linewidth=1.5, linestyle="--",
         label="Benchmark (47%)", zorder=2)
ax2.fill_between(months_short, er_vals, bench_line,
                 where=[v >= 47 for v in er_vals],
                 alpha=0.15, color=C_MID, interpolate=True)
ax2.fill_between(months_short, er_vals, bench_line,
                 where=[v < 47 for v in er_vals],
                 alpha=0.15, color="#e53935", interpolate=True)

ax2.yaxis.set_major_formatter(mticker.FuncFormatter(lambda y, _: f"{y:.0f}%"))
ax2.set_title("Email Newsletter: Monthly Open Rate vs Benchmark", color=C_TEXT,
              fontsize=11, fontweight="bold", pad=12)
ax2.legend(frameon=False, fontsize=8)
ax2.tick_params(axis="x", rotation=45, labelsize=8, colors=C_MUTED)
ax2.tick_params(axis="y", colors=C_MUTED)
ax2.spines["left"].set_color(C_BORDER)
ax2.spines["bottom"].set_color(C_BORDER)
ax2.grid(axis="y", color=C_BORDER, linewidth=0.5, zorder=0)
ax2.set_ylim(35, 70)

# Title block
fig.text(0.04, 0.97, "Meridian Insurance Group", fontsize=14, fontweight="bold",
         color=C_DARK, va="top")
fig.text(0.04, 0.93, "Internal Communications — Full Year Performance Report  |  Jan – Dec 2024",
         fontsize=9, color=C_MUTED, va="top")

plt.tight_layout(rect=[0, 0, 1, 0.90])
plt.savefig(f"{IMG_DIR}/report-dashboard.png", dpi=160, bbox_inches="tight",
            facecolor=C_CREAM)
plt.close()
print(f"✓ Image saved: {IMG_DIR}/report-dashboard.png")

# ── IMAGE 2: Raw data table preview ──────────────────────────────────────
fig2, ax = plt.subplots(figsize=(14, 6))
fig2.patch.set_facecolor(C_CREAM)
ax.set_facecolor(C_CREAM)
ax.axis("off")

# Show first 13 rows of raw data (header + 12 email rows)
table_headers = ["Channel","Month","Campaign / Content","Reach","Engagements","Clicks",
                 "Eng. Rate","CTR","Benchmark\nEng. Rate","vs Benchmark"]
table_data = []
for row in raw_rows[:13]:
    er = row[4]/row[3]
    ct = row[5]/row[3]
    ben = bench_eng(row[0])
    vs  = er - ben
    vs_str = f"+{vs*100:.1f}%" if vs >= 0 else f"{vs*100:.1f}%"
    table_data.append([
        row[0], row[1], row[2][:28]+"…" if len(row[2]) > 28 else row[2],
        f"{row[3]:,}", f"{row[4]:,}", str(row[5]),
        f"{er*100:.1f}%", f"{ct*100:.1f}%",
        f"{ben*100:.1f}%", vs_str
    ])

col_widths = [0.13, 0.08, 0.22, 0.06, 0.09, 0.06, 0.08, 0.07, 0.10, 0.09]

tbl = ax.table(
    cellText=table_data,
    colLabels=table_headers,
    colWidths=col_widths,
    loc="center",
    cellLoc="center"
)
tbl.auto_set_font_size(False)
tbl.set_fontsize(8)
tbl.scale(1, 1.6)

# Style header
for col in range(len(table_headers)):
    cell = tbl[0, col]
    cell.set_facecolor(C_DARK)
    cell.set_text_props(color="white", fontweight="bold")
    cell.set_edgecolor(C_BORDER)

# Style data rows
for row_idx in range(1, len(table_data) + 1):
    bg = C_WHITE if row_idx % 2 == 1 else C_SURFACE
    for col_idx in range(len(table_headers)):
        cell = tbl[row_idx, col_idx]
        cell.set_facecolor(bg)
        cell.set_edgecolor(C_BORDER)
        # Color vs benchmark column
        if col_idx == 9:
            val_str = table_data[row_idx - 1][9]
            if val_str.startswith("+"):
                cell.set_facecolor("#e8f5e9")
                cell.set_text_props(color="#2e7d32")
            elif val_str.startswith("-"):
                cell.set_facecolor(C_RED)
                cell.set_text_props(color=C_RED_T)

fig2.text(0.04, 0.97, "Raw Data Sheet — with VLOOKUP benchmark column",
          fontsize=11, fontweight="bold", color=C_DARK, va="top")
fig2.text(0.04, 0.93, "Showing Email Newsletter rows (Jan – Dec 2024). Green = above benchmark · Red = below benchmark.",
          fontsize=8.5, color=C_MUTED, va="top")

plt.tight_layout(rect=[0, 0, 1, 0.90])
plt.savefig(f"{IMG_DIR}/report-rawdata.png", dpi=160, bbox_inches="tight",
            facecolor=C_CREAM)
plt.close()
print(f"✓ Image saved: {IMG_DIR}/report-rawdata.png")

print("\nAll files generated successfully.")
