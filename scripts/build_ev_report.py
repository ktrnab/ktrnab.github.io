"""
Build EV Charging Network — Operational Intelligence Report
Generates:
  - assets/meridian-ev-charging-report-2023.xlsx
  - assets/images/work/report-dashboard.png
  - assets/images/work/report-rawdata.png
"""

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                             GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import random, math, os

os.chdir('/Users/katrinab/ktrnab.github.io')
random.seed(42)

FNAME = 'assets/meridian-ev-charging-report-2023.xlsx'
wb = openpyxl.Workbook()

# ── Palette ─────────────────────────────────────────────────────────────────
GREEN_DARK  = '1A4731'
GREEN_MID   = '2D6A4F'
GREEN_LIGHT = 'B7E4C7'
AMBER       = 'F4A261'
RED_LIGHT   = 'FFD6D6'
BLUE_LIGHT  = 'D6EAF8'
PURPLE_LIGHT= 'E8DAEF'
GREY_LIGHT  = 'F2F2F2'
WHITE       = 'FFFFFF'
DARK_TEXT   = '1A1A1A'

def hfill(color): return PatternFill('solid', fgColor=color)
def font(bold=False, color=DARK_TEXT, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left():   return Alignment(horizontal='left',   vertical='center', wrap_text=True)
def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

MONTHS = ['Jan','Feb','Mar','Apr','May','Jun',
          'Jul','Aug','Sep','Oct','Nov','Dec']
STATION_TYPES = ['Level 2 (L2)', 'DC Fast Charger (DCFC)', 'Level 1 (L1)']
# Target utilisation rates per type
TARGETS = {'Level 2 (L2)': 0.42, 'DC Fast Charger (DCFC)': 0.65, 'Level 1 (L1)': 0.28}

# Monthly utilisation — EV shows winter dip (cold = slower charging, less range anxiety trips)
base_util = {
    'Level 2 (L2)':          [.38,.36,.40,.44,.48,.50,.52,.51,.49,.44,.40,.36],
    'DC Fast Charger (DCFC)':[.60,.57,.63,.66,.70,.72,.74,.73,.70,.67,.62,.58],
    'Level 1 (L1)':          [.24,.22,.25,.28,.30,.31,.32,.30,.28,.26,.23,.21],
}
# Avg kWh per session
avg_kwh = {
    'Level 2 (L2)':          [18.2,17.8,19.1,20.4,21.0,21.5,21.8,21.3,20.6,19.5,18.0,17.2],
    'DC Fast Charger (DCFC)':[38.5,36.9,40.2,42.1,43.8,44.5,45.2,44.8,43.0,41.2,38.8,36.4],
    'Level 1 (L1)':          [7.8, 7.2, 8.1, 8.9, 9.2, 9.5, 9.6, 9.3, 8.8, 8.2, 7.5, 7.0],
}
# Sessions per month per type
sessions = {
    'Level 2 (L2)':          [312,285,340,388,421,445,458,442,415,374,325,298],
    'DC Fast Charger (DCFC)':[148,132,159,178,196,208,215,210,198,183,152,138],
    'Level 1 (L1)':          [87, 79, 92, 105,114,120,124,118,109,96, 82, 75],
}

# ── Sheet 1: Rate Table (VLOOKUP lookup) ────────────────────────────────────
ws_rates = wb.active
ws_rates.title = 'Rate Table'
ws_rates.sheet_view.showGridLines = False

ws_rates['A1'] = 'EV Charging — Time-of-Use Rate Table'
ws_rates['A1'].font = font(bold=True, size=13, color=WHITE)
ws_rates['A1'].fill = hfill(GREEN_DARK)
ws_rates.merge_cells('A1:D1')
ws_rates['A1'].alignment = center()
ws_rates.row_dimensions[1].height = 28

headers = ['Period', 'Start Hour (0–23)', 'Rate ($/kWh)', 'Rate Band']
for col, h in enumerate(headers, 1):
    c = ws_rates.cell(2, col, h)
    c.font = font(bold=True, color=WHITE, size=9)
    c.fill = hfill(GREEN_MID)
    c.alignment = center()
    c.border = thin_border()

rate_data = [
    ('Off-Peak',  0,  0.082, 'Low'),
    ('Off-Peak',  1,  0.082, 'Low'),
    ('Off-Peak',  2,  0.082, 'Low'),
    ('Off-Peak',  3,  0.082, 'Low'),
    ('Off-Peak',  4,  0.082, 'Low'),
    ('Off-Peak',  5,  0.082, 'Low'),
    ('Off-Peak',  6,  0.082, 'Low'),
    ('Mid-Peak',  7,  0.142, 'Medium'),
    ('Mid-Peak',  8,  0.142, 'Medium'),
    ('Mid-Peak',  9,  0.142, 'Medium'),
    ('Mid-Peak', 10,  0.142, 'Medium'),
    ('Mid-Peak', 11,  0.142, 'Medium'),
    ('Mid-Peak', 12,  0.142, 'Medium'),
    ('Mid-Peak', 13,  0.142, 'Medium'),
    ('On-Peak',  14,  0.218, 'High'),
    ('On-Peak',  15,  0.218, 'High'),
    ('On-Peak',  16,  0.218, 'High'),
    ('On-Peak',  17,  0.218, 'High'),
    ('On-Peak',  18,  0.218, 'High'),
    ('On-Peak',  19,  0.218, 'High'),
    ('On-Peak',  20,  0.218, 'High'),
    ('On-Peak',  21,  0.218, 'High'),
    ('Mid-Peak', 22,  0.142, 'Medium'),
    ('Off-Peak', 23,  0.082, 'Low'),
]
band_colors = {
    'Low': BLUE_LIGHT, 'Medium': 'FFF3CD', 'High': RED_LIGHT,
    'Off-Peak': BLUE_LIGHT, 'Mid-Peak': 'FFF3CD', 'On-Peak': RED_LIGHT,
}
for row_i, row in enumerate(rate_data, 3):
    color = band_colors[row[3]]
    for col_i, val in enumerate(row, 1):
        c = ws_rates.cell(row_i, col_i, val)
        c.fill = hfill(color)
        c.alignment = center()
        c.border = thin_border()
        c.font = font(size=9)
        if col_i == 3:
            c.number_format = '$0.000'

ws_rates.column_dimensions['A'].width = 14
ws_rates.column_dimensions['B'].width = 20
ws_rates.column_dimensions['C'].width = 16
ws_rates.column_dimensions['D'].width = 14

# ── Sheet 2: Station Specs ───────────────────────────────────────────────────
ws_specs = wb.create_sheet('Station Specs')
ws_specs.sheet_view.showGridLines = False

ws_specs['A1'] = 'Station Hardware Specifications'
ws_specs['A1'].font = font(bold=True, size=13, color=WHITE)
ws_specs['A1'].fill = hfill(GREEN_DARK)
ws_specs.merge_cells('A1:F1')
ws_specs['A1'].alignment = center()
ws_specs.row_dimensions[1].height = 28

spec_headers = ['Station Type', 'Max Power (kW)', 'Avg Session (hrs)', 'Connector Type', 'Target Utilisation', 'Notes']
for col, h in enumerate(spec_headers, 1):
    c = ws_specs.cell(2, col, h)
    c.font = font(bold=True, color=WHITE, size=9)
    c.fill = hfill(GREEN_MID)
    c.alignment = center()
    c.border = thin_border()

spec_data = [
    ('Level 1 (L1)',          1.4,  8.5, 'Standard 120V',    '28%', 'Overnight residential; lowest throughput'),
    ('Level 2 (L2)',          7.2,  2.8, 'J1772 / Type 2',   '42%', 'Primary fleet & workplace charger'),
    ('DC Fast Charger (DCFC)',150,  0.4, 'CCS / CHAdeMO',    '65%', 'Highway corridor; highest demand & revenue'),
]
row_colors = [BLUE_LIGHT, GREEN_LIGHT, 'FFF3CD']
for row_i, (row, color) in enumerate(zip(spec_data, row_colors), 3):
    for col_i, val in enumerate(row, 1):
        c = ws_specs.cell(row_i, col_i, val)
        c.fill = hfill(color)
        c.alignment = center()
        c.border = thin_border()
        c.font = font(size=9)

for col, w in zip('ABCDEF', [22,18,22,20,20,36]):
    ws_specs.column_dimensions[col].width = w

# ── Sheet 3: Raw Data ────────────────────────────────────────────────────────
ws_raw = wb.create_sheet('Raw Data')
ws_raw.sheet_view.showGridLines = False

ws_raw['A1'] = 'Monthly Charging Session Data — 2023'
ws_raw['A1'].font = font(bold=True, size=13, color=WHITE)
ws_raw['A1'].fill = hfill(GREEN_DARK)
ws_raw.merge_cells('A1:J1')
ws_raw['A1'].alignment = center()
ws_raw.row_dimensions[1].height = 28

raw_headers = [
    'Station Type', 'Month', 'Sessions',
    'Avg kWh / Session', 'Total kWh Delivered',
    'Utilisation Rate', 'Target Util.',
    'vs Target (pp)',
    'Rate ($/kWh)',   # VLOOKUP from Rate Table (peak hour = 17 for illustration)
    'Est. Revenue ($)'
]
for col, h in enumerate(raw_headers, 1):
    c = ws_raw.cell(2, col, h)
    c.font = font(bold=True, color=WHITE, size=8)
    c.fill = hfill(GREEN_MID)
    c.alignment = center()
    c.border = thin_border()

row_num = 3
for stype in STATION_TYPES:
    for m_i, month in enumerate(MONTHS):
        util  = base_util[stype][m_i]
        kwh   = avg_kwh[stype][m_i]
        sess  = sessions[stype][m_i]
        total = round(sess * kwh, 1)
        target = TARGETS[stype]
        vs_tgt = round((util - target) * 100, 1)  # percentage points

        # VLOOKUP formula: lookup hour 17 (on-peak) from Rate Table col C
        vlookup_formula = '=VLOOKUP(17,\'Rate Table\'!$B$3:$C$26,2,TRUE)'
        revenue_formula = f'=C{row_num}*D{row_num}*I{row_num}'

        row_vals = [
            stype, month, sess,
            kwh, total,
            util, target,
            vs_tgt,
            vlookup_formula,
            revenue_formula
        ]
        alt = (row_num % 2 == 0)
        for col_i, val in enumerate(row_vals, 1):
            c = ws_raw.cell(row_num, col_i, val)
            c.border = thin_border()
            c.font = font(size=8)
            c.alignment = center()
            if col_i not in (9, 10):  # leave formula cells plain
                c.fill = hfill(GREY_LIGHT if alt else WHITE)
            if col_i == 6:
                c.number_format = '0.0%'
            if col_i == 7:
                c.number_format = '0.0%'
            if col_i == 8:
                c.number_format = '+0.0;-0.0;0.0'
            if col_i == 9:
                c.number_format = '$0.000'
            if col_i == 10:
                c.number_format = '$#,##0'
        row_num += 1

# Conditional formatting: vs Target column (H) green/red
last_row = row_num - 1
ws_raw.conditional_formatting.add(
    f'H3:H{last_row}',
    CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill('solid', fgColor='C8E6C9'))
)
ws_raw.conditional_formatting.add(
    f'H3:H{last_row}',
    CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill('solid', fgColor='FFCDD2'))
)

col_widths = [22, 8, 10, 18, 20, 16, 14, 16, 14, 16]
for col, w in enumerate(col_widths, 1):
    ws_raw.column_dimensions[get_column_letter(col)].width = w

# ── Sheet 4: Duck Curve (hourly load pivot) ──────────────────────────────────
ws_duck = wb.create_sheet('Duck Curve')
ws_duck.sheet_view.showGridLines = False

ws_duck['A1'] = 'Hourly Charging Load Profile — "Duck Curve" (Annual Average)'
ws_duck['A1'].font = font(bold=True, size=13, color=WHITE)
ws_duck['A1'].fill = hfill(GREEN_DARK)
ws_duck.merge_cells('A1:D1')
ws_duck['A1'].alignment = center()
ws_duck.row_dimensions[1].height = 28

duck_headers = ['Hour of Day', 'Label', 'Avg Grid Load (kW)', 'Rate Band']
for col, h in enumerate(duck_headers, 1):
    c = ws_duck.cell(2, col, h)
    c.font = font(bold=True, color=WHITE, size=9)
    c.fill = hfill(GREEN_MID)
    c.alignment = center()
    c.border = thin_border()

# Duck curve shape: low overnight, dip midday (solar), spike 5-9pm
duck_load = [
    42, 38, 34, 31, 30, 33,   # 0-5 (off-peak overnight)
    48, 72, 91, 98, 94, 88,   # 6-11 (morning ramp)
    82, 76, 71, 78, 98, 134,  # 12-17 (midday dip, then evening surge)
    158,167,152,131,108, 82,  # 18-23 (peak then wind-down)
]
rate_band_by_hour = (
    ['Off-Peak']*7 + ['Mid-Peak']*7 + ['On-Peak']*8 + ['Mid-Peak']*1 + ['Off-Peak']*1
)
for hour, (load, band) in enumerate(zip(duck_load, rate_band_by_hour)):
    r = hour + 3
    color = band_colors[band]
    label = f'{hour:02d}:00'
    for col_i, val in enumerate([hour, label, load, band], 1):
        c = ws_duck.cell(r, col_i, val)
        c.fill = hfill(color)
        c.alignment = center()
        c.border = thin_border()
        c.font = font(size=9)

for col, w in zip('ABCD', [14, 12, 22, 14]):
    ws_duck.column_dimensions[col].width = w

# ── Sheet 5: Station Summary (Pivot-style) ──────────────────────────────────
ws_sum = wb.create_sheet('Station Summary')
ws_sum.sheet_view.showGridLines = False

ws_sum['A1'] = 'Annual Performance Summary by Station Type'
ws_sum['A1'].font = font(bold=True, size=13, color=WHITE)
ws_sum['A1'].fill = hfill(GREEN_DARK)
ws_sum.merge_cells('A1:G1')
ws_sum['A1'].alignment = center()
ws_sum.row_dimensions[1].height = 28

sum_headers = ['Station Type', 'Total Sessions', 'Total kWh', 'Avg Utilisation',
               'Target Util.', 'vs Target (pp)', 'Est. Annual Revenue ($)']
for col, h in enumerate(sum_headers, 1):
    c = ws_sum.cell(2, col, h)
    c.font = font(bold=True, color=WHITE, size=9)
    c.fill = hfill(GREEN_MID)
    c.alignment = center()
    c.border = thin_border()

row_colors_sum = [BLUE_LIGHT, GREEN_LIGHT, 'FFF3CD']
for row_i, (stype, color) in enumerate(zip(STATION_TYPES, row_colors_sum), 3):
    total_sess = sum(sessions[stype])
    total_kwh  = sum(s * k for s, k in zip(sessions[stype], avg_kwh[stype]))
    avg_util   = sum(base_util[stype]) / 12
    target     = TARGETS[stype]
    vs_tgt     = round((avg_util - target) * 100, 1)
    revenue    = round(total_kwh * 0.218, 0)  # on-peak rate for worst-case

    row_vals = [stype, total_sess, round(total_kwh), f'{avg_util:.1%}',
                f'{target:.0%}', vs_tgt, revenue]
    for col_i, val in enumerate(row_vals, 1):
        c = ws_sum.cell(row_i, col_i, val)
        c.fill = hfill(color)
        c.alignment = center()
        c.border = thin_border()
        c.font = font(size=9, bold=(col_i == 1))
        if col_i == 7:
            c.number_format = '$#,##0'

for col, w in zip('ABCDEFG', [22, 16, 14, 18, 14, 18, 22]):
    ws_sum.column_dimensions[col].width = w

# ── Sheet 6: Dashboard ───────────────────────────────────────────────────────
ws_dash = wb.create_sheet('Dashboard')
ws_dash.sheet_view.showGridLines = False
# Move to first position
wb.move_sheet('Dashboard', offset=-(len(wb.sheetnames)-1))

ws_dash['A1'] = 'EV Charging Network — Operational Intelligence Dashboard'
ws_dash['A1'].font = font(bold=True, size=14, color=WHITE)
ws_dash['A1'].fill = hfill(GREEN_DARK)
ws_dash.merge_cells('A1:K1')
ws_dash['A1'].alignment = center()
ws_dash.row_dimensions[1].height = 32

ws_dash['A2'] = 'GreenGrid EV Solutions  ·  FY 2023  ·  Annual Network Review'
ws_dash['A2'].font = font(italic=True, size=9, color='555555')
ws_dash.merge_cells('A2:K2')
ws_dash['A2'].alignment = center()

# KPI cards row
kpi_labels = [
    'Total Energy Delivered', 'Total Sessions', 'Avg DCFC Utilisation',
    'Peak Load Hour', 'Est. Annual Revenue'
]
# Compute real values
total_kwh_all = sum(
    sessions[s][m] * avg_kwh[s][m]
    for s in STATION_TYPES for m in range(12)
)
total_sess_all = sum(sessions[s][m] for s in STATION_TYPES for m in range(12))
dcfc_util = sum(base_util['DC Fast Charger (DCFC)']) / 12
revenue_all = total_kwh_all * 0.218

kpi_values = [
    f'{total_kwh_all/1000:.1f} MWh',
    f'{total_sess_all:,}',
    f'{dcfc_util:.0%}',
    '18:00 – 20:00',
    f'${revenue_all:,.0f}'
]
kpi_cols = [1, 3, 5, 7, 9]  # A, C, E, G, I (merged pairs)
kpi_colors = [GREEN_LIGHT, BLUE_LIGHT, 'FFF3CD', 'FDEBD0', PURPLE_LIGHT]

ws_dash.row_dimensions[4].height = 18
ws_dash.row_dimensions[5].height = 32
ws_dash.row_dimensions[6].height = 20

for (label, value, col, color) in zip(kpi_labels, kpi_values, kpi_cols, kpi_colors):
    end_col = col + 1
    lc = get_column_letter(col)
    ec = get_column_letter(end_col)
    ws_dash.merge_cells(f'{lc}4:{ec}4')
    ws_dash.merge_cells(f'{lc}5:{ec}5')
    ws_dash.merge_cells(f'{lc}6:{ec}6')
    c_label = ws_dash[f'{lc}4']
    c_label.value = label
    c_label.font = font(size=8, color='666666')
    c_label.alignment = center()
    c_label.fill = hfill(color)
    c_val = ws_dash[f'{lc}5']
    c_val.value = value
    c_val.font = font(bold=True, size=16, color=GREEN_DARK)
    c_val.alignment = center()
    c_val.fill = hfill(color)
    c_sub = ws_dash[f'{lc}6']
    c_sub.value = '2023 Total'
    c_sub.font = font(size=7, color='888888', italic=True)
    c_sub.alignment = center()
    c_sub.fill = hfill(color)

# Spacer
ws_dash.row_dimensions[7].height = 8

# Utilisation data for bar chart
ws_dash['A8'] = 'Station Type'
ws_dash['B8'] = 'Avg Utilisation'
ws_dash['C8'] = 'Target'
for col in 'ABC':
    c = ws_dash[f'{col}8']
    c.font = font(bold=True, size=9, color=WHITE)
    c.fill = hfill(GREEN_MID)
    c.alignment = center()

chart_labels = ['Level 1', 'Level 2', 'DC Fast']
chart_util   = [
    sum(base_util['Level 1 (L1)'])/12,
    sum(base_util['Level 2 (L2)'])/12,
    sum(base_util['DC Fast Charger (DCFC)'])/12,
]
chart_targets = [
    TARGETS['Level 1 (L1)'],
    TARGETS['Level 2 (L2)'],
    TARGETS['DC Fast Charger (DCFC)'],
]
for i, (label, util, tgt) in enumerate(zip(chart_labels, chart_util, chart_targets), 9):
    ws_dash[f'A{i}'] = label
    ws_dash[f'B{i}'] = round(util, 3)
    ws_dash[f'C{i}'] = tgt
    ws_dash[f'B{i}'].number_format = '0%'
    ws_dash[f'C{i}'].number_format = '0%'

# Monthly DCFC utilisation for line chart
ws_dash['E8'] = 'Month'
ws_dash['F8'] = 'DCFC Util.'
ws_dash['G8'] = 'Target (65%)'
for col in 'EFG':
    c = ws_dash[f'{col}8']
    c.font = font(bold=True, size=9, color=WHITE)
    c.fill = hfill(GREEN_MID)
    c.alignment = center()

for i, (month, util) in enumerate(zip(MONTHS, base_util['DC Fast Charger (DCFC)']), 9):
    ws_dash[f'E{i}'] = month
    ws_dash[f'F{i}'] = util
    ws_dash[f'G{i}'] = 0.65
    ws_dash[f'F{i}'].number_format = '0%'
    ws_dash[f'G{i}'].number_format = '0%'

wb.save(FNAME)
print(f"Saved {FNAME}")


# ── PNG Images ───────────────────────────────────────────────────────────────
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.patches import FancyBboxPatch
import numpy as np

W, H = 14, 8.5
DPI = 140

import matplotlib
matplotlib.rcParams['font.family'] = 'serif'
matplotlib.rcParams['font.serif'] = ['DejaVu Serif', 'Georgia', 'Times New Roman', 'serif']

def fig(facecolor='#FFFFFF'):
    f, ax = plt.subplots(figsize=(W, H), facecolor=facecolor)
    ax.set_xlim(0, W)
    ax.set_ylim(0, H)
    ax.axis('off')
    return f, ax

def rect(ax, x, y, w, h, color, zorder=2, radius=0, alpha=1, ec='none', lw=1):
    if radius:
        p = FancyBboxPatch((x, y), w, h,
                           boxstyle=f"round,pad=0,rounding_size={radius}",
                           facecolor=color, edgecolor=ec, linewidth=lw,
                           zorder=zorder, alpha=alpha)
    else:
        p = patches.Rectangle((x, y), w, h,
                               facecolor=color, edgecolor=ec, linewidth=lw,
                               zorder=zorder, alpha=alpha)
    ax.add_patch(p)

def txt(ax, x, y, s, size=9, color='#333333', ha='left', va='bottom',
        bold=False, zorder=20):
    ax.text(x, y, s, fontsize=size, color=color, ha=ha, va=va,
            fontweight='bold' if bold else 'normal', zorder=zorder)

# ── Dashboard PNG ────────────────────────────────────────────────────────────
def build_dashboard():
    import matplotlib.patheffects as pe
    from matplotlib.patches import Patch

    BG       = '#080D1A'
    CARD_BG  = '#0D1628'
    GRID_COL = '#111E30'
    NEON     = '#00E5A0'
    CYAN     = '#00C2D4'
    AMBER    = '#FFB700'
    PINK     = '#FF4D6D'
    VIOLET   = '#B388FF'
    TEXT_HI  = '#D8EEF8'
    TEXT_MID = '#6B8CAE'
    TEXT_DIM = '#2E4460'
    SPINE    = '#1A2E48'

    f, ax = fig(BG)

    # Subtle background grid
    for xi in np.arange(0, W + 1, 1.0):
        ax.axvline(xi, color=GRID_COL, linewidth=0.4, zorder=1)
    for yi in np.arange(0, H + 1, 1.0):
        ax.axhline(yi, color=GRID_COL, linewidth=0.4, zorder=1)

    # Header
    rect(ax, 0, H - 0.72, W, 0.72, '#09122A')
    rect(ax, 0, H - 0.75, W, 0.035, NEON, zorder=5)   # neon accent line
    txt(ax, W/2, H - 0.42, 'EV Charging Network — Operational Intelligence Dashboard',
        size=13, color=TEXT_HI, ha='center', bold=True, zorder=10)
    txt(ax, W/2, H - 0.66, 'BayWa Mobility Solutions  ·  FY 2023  ·  Operational Review',
        size=7.5, color=TEXT_MID, ha='center', zorder=10)

    # KPI cards
    kpi_data = [
        ('Total Energy',      f'{total_kwh_all/1000:.1f} MWh', NEON),
        ('Total Sessions',    f'{total_sess_all:,}',            CYAN),
        ('DCFC Utilisation',  f'{dcfc_util:.0%}',               AMBER),
        ('Peak Window',       '18:00–20:00',                    PINK),
        ('Est. Revenue',      f'${revenue_all:,.0f}',           VIOLET),
    ]
    card_w = (W - 0.8) / 5 - 0.12
    for i, (label, val, accent) in enumerate(kpi_data):
        cx = 0.4 + i * (card_w + 0.12)
        cy = H - 1.78
        # Outer glow
        rect(ax, cx - 0.06, cy - 0.06, card_w + 0.12, 0.92, accent,
             radius=0.13, alpha=0.12, zorder=3)
        # Card body
        rect(ax, cx, cy, card_w, 0.80, CARD_BG, radius=0.1,
             ec=accent, lw=0.8, alpha=1, zorder=4)
        # Top accent stripe
        rect(ax, cx + 0.02, cy + 0.76, card_w - 0.04, 0.035, accent,
             radius=0.04, zorder=5)
        txt(ax, cx + card_w/2, cy + 0.44, val,
            size=15, color=accent, ha='center', bold=True, zorder=6)
        txt(ax, cx + card_w/2, cy + 0.10, label,
            size=6.5, color=TEXT_MID, ha='center', zorder=6)

    # ── Left chart: Utilisation vs Target ────────────────────────────────────
    ax_bar = f.add_axes([0.04, 0.12, 0.42, 0.51])
    ax_bar.set_facecolor('#0A1422')

    station_labels = ['Level 2\n(L2)', 'DC Fast\n(DCFC)', 'Level 1\n(L1)']
    util_vals   = [sum(base_util[s])/12 for s in STATION_TYPES]
    target_vals = [TARGETS[s] for s in STATION_TYPES]

    x = np.arange(len(station_labels))
    w = 0.32

    # Glow layer
    ax_bar.bar(x - w/2, util_vals, w * 1.8, color=NEON, alpha=0.08, zorder=2)
    bars1 = ax_bar.bar(x - w/2, util_vals, w, color=NEON, alpha=0.85, zorder=3)
    ax_bar.bar(x + w/2, target_vals, w, color='none',
               edgecolor=AMBER, linewidth=1.5, linestyle='--', zorder=3)

    ax_bar.set_xticks(x)
    ax_bar.set_xticklabels(station_labels, fontsize=8, color=TEXT_MID)
    ax_bar.set_ylabel('Utilisation Rate', fontsize=8, color=TEXT_MID)
    ax_bar.set_ylim(0, 0.85)
    ax_bar.yaxis.set_major_formatter(plt.FuncFormatter(lambda v, _: f'{v:.0%}'))
    ax_bar.tick_params(colors=TEXT_DIM)
    ax_bar.set_title('Station Utilisation vs Target', fontsize=10, fontweight='bold',
                     color=TEXT_HI, pad=8)
    ax_bar.legend(['Avg Utilisation', 'Target'], fontsize=7,
                  framealpha=0.15, facecolor=CARD_BG, edgecolor=SPINE,
                  labelcolor=TEXT_MID)
    ax_bar.grid(axis='y', alpha=0.12, color='#1E3050', zorder=0)
    for spine in ax_bar.spines.values():
        spine.set_edgecolor(SPINE)
    ax_bar.spines[['top','right']].set_visible(False)

    for bar in bars1:
        h_val = bar.get_height()
        t = ax_bar.text(bar.get_x() + bar.get_width()/2, h_val + 0.012,
                        f'{h_val:.0%}', ha='center', va='bottom',
                        fontsize=8, fontweight='bold', color=NEON)
        t.set_path_effects([pe.withStroke(linewidth=3, foreground=NEON, alpha=0.25)])

    # ── Right chart: Duck Curve ───────────────────────────────────────────────
    ax_duck = f.add_axes([0.55, 0.12, 0.42, 0.51])
    ax_duck.set_facecolor('#0A1422')

    hours = list(range(24))
    off_peak_hours = list(range(7)) + [23]
    mid_peak_hours = list(range(7, 14)) + [22]
    on_peak_hours  = list(range(14, 22))

    bar_colors = []
    for h in hours:
        if h in on_peak_hours:   bar_colors.append(PINK)
        elif h in mid_peak_hours: bar_colors.append(AMBER)
        else:                     bar_colors.append(CYAN)

    ax_duck.bar(hours, duck_load, color=bar_colors, edgecolor='none', alpha=0.35, zorder=3)

    # Glowing curve — layered
    ax_duck.plot(hours, duck_load, color=CYAN, linewidth=7, alpha=0.10, zorder=4)
    ax_duck.plot(hours, duck_load, color=CYAN, linewidth=3, alpha=0.22, zorder=5)
    ax_duck.plot(hours, duck_load, color=CYAN, linewidth=1.5, zorder=6,
                 marker='o', markersize=3,
                 markerfacecolor=CYAN, markeredgecolor='none')

    ax_duck.axvspan(14, 21, alpha=0.07, color=PINK, zorder=2)
    ax_duck.text(17.5, 175, 'On-Peak\n(TOU High Rate)', ha='center', fontsize=7,
                 color=PINK, style='italic')

    ax_duck.set_xticks(range(0, 24, 3))
    ax_duck.set_xticklabels([f'{h:02d}:00' for h in range(0, 24, 3)],
                             fontsize=7, color=TEXT_MID)
    ax_duck.set_ylabel('Avg Grid Load (kW)', fontsize=8, color=TEXT_MID)
    ax_duck.set_ylim(0, 200)
    ax_duck.set_title('Hourly Load Profile — "Duck Curve"', fontsize=10,
                      fontweight='bold', color=TEXT_HI, pad=8)
    ax_duck.grid(axis='y', alpha=0.12, color='#1E3050', zorder=0)
    ax_duck.tick_params(colors=TEXT_DIM)
    for spine in ax_duck.spines.values():
        spine.set_edgecolor(SPINE)
    ax_duck.spines[['top','right']].set_visible(False)

    legend_elements = [
        Patch(facecolor=CYAN,  alpha=0.6, label='Off-Peak ($0.082)'),
        Patch(facecolor=AMBER, alpha=0.6, label='Mid-Peak ($0.142)'),
        Patch(facecolor=PINK,  alpha=0.6, label='On-Peak ($0.218)'),
    ]
    ax_duck.legend(handles=legend_elements, fontsize=7, loc='upper left',
                   framealpha=0.15, facecolor=CARD_BG, edgecolor=SPINE,
                   labelcolor=TEXT_MID)

    f.savefig('assets/images/work/report-dashboard.png', dpi=DPI,
              bbox_inches='tight', facecolor=BG)
    plt.close(f)
    print("Saved report-dashboard.png")


# ── Raw Data PNG ─────────────────────────────────────────────────────────────
def build_rawdata():
    f, ax = fig()

    rect(ax, 0, 0, W, H, '#FFFFFF')

    # Header
    rect(ax, 0, H-0.55, W, 0.55, '#1A4731')
    txt(ax, W/2, H-0.3, 'Monthly Charging Session Data — Raw Data Sheet with VLOOKUP',
        size=11, color='#FFFFFF', ha='center', bold=True)

    # 9 columns — table occupies x=0.05..11.2, callout lives in right margin
    col_headers = ['Station Type','Month','Sessions','Avg kWh',
                   'Utilisation','Target','vs Target','Rate ($/kWh)','Revenue']
    col_xs = [0.12, 2.25, 3.1, 4.0, 5.0, 5.95, 6.9, 8.05, 9.4]
    col_ws = [2.0,  0.75, 0.8, 0.85, 0.85, 0.85, 1.0,  1.2,  1.5]

    header_y = H - 0.9
    rect(ax, 0.05, header_y - 0.1, 11.0, 0.44, '#2D6A4F')
    for hdr, cx in zip(col_headers, col_xs):
        txt(ax, cx, header_y + 0.07, hdr, size=7, color='#FFFFFF', ha='left', bold=True)

    # TOU rate lookup — varies by station type to illustrate VLOOKUP concept
    tou_rates = {
        'Level 2 (L2)':          [0.082]*6 + [0.142]*4 + [0.082]*2,   # mostly off/mid-peak
        'DC Fast Charger (DCFC)':[0.218]*12,                            # peak evening sessions
        'Level 1 (L1)':          [0.082]*12,                            # overnight off-peak
    }

    sample_rows = []
    for stype in STATION_TYPES:
        for m_i, month in enumerate(MONTHS):
            util   = base_util[stype][m_i]
            kwh    = avg_kwh[stype][m_i]
            sess   = sessions[stype][m_i]
            target = TARGETS[stype]
            vs_tgt = round((util - target) * 100, 1)
            rate   = tou_rates[stype][m_i]
            rev    = sess * kwh * rate
            sample_rows.append((
                stype[:12], month, sess,
                f'{kwh:.1f}',
                f'{util:.1%}', f'{target:.0%}',
                vs_tgt, f'${rate:.3f}', f'${rev:,.0f}'
            ))

    shown = sample_rows[:18]
    row_h = 0.38
    start_y = header_y - 0.15
    for i, row in enumerate(shown):
        ry = start_y - (i + 1) * row_h
        if ry < 0.15: break
        bg = '#F8F9FA' if i % 2 == 0 else '#FFFFFF'

        vs_tgt_val = row[6]
        vs_color   = '#C8E6C9' if vs_tgt_val >= 0 else '#FFCDD2'
        rate_color = '#FFF8E1'

        rect(ax, 0.05, ry - 0.03, 11.0, row_h - 0.04, bg)

        for j, (val, cx) in enumerate(zip(row, col_xs)):
            cell_color = None
            if j == 6: cell_color = vs_color
            if j == 7: cell_color = rate_color
            if cell_color:
                rect(ax, cx - 0.05, ry - 0.03,
                     col_ws[j] + 0.02, row_h - 0.04, cell_color, radius=0.03)

            display = f'+{val:.1f}pp' if j == 6 and isinstance(val, float) and val >= 0 \
                      else (f'{val:.1f}pp' if j == 6 and isinstance(val, float) else str(val))
            color = '#1A4731' if j == 6 and isinstance(val, float) and val >= 0 \
                    else ('#CC0000' if j == 6 and isinstance(val, float) else '#333333')
            txt(ax, cx, ry + 0.12, display, size=7.5, color=color,
                bold=(j == 0 or j == 6))

        ax.axhline(ry - 0.04, color='#EEEEEE', linewidth=0.5, zorder=3)

    # VLOOKUP callout — right margin, clear of table
    cx, cy, cw, ch = 11.3, 2.8, 2.55, 5.0
    rect(ax, cx, cy, cw, ch, '#FFF8E1', radius=0.12, ec='#F4A261', lw=1.5, zorder=10)
    txt(ax, cx + 0.18, cy + ch - 0.35, 'VLOOKUP Formula', size=8,
        color='#7D4E00', bold=True, zorder=11)
    for dy, line, bold, color in [
        (0.75, '=VLOOKUP(', False, '#333333'),
        (1.08, '  SessionHour,', False, '#333333'),
        (1.38, "  'Rate Table'", True,  '#1A4731'),
        (1.62, '  !B3:C26, 2, TRUE)', True,  '#1A4731'),
        (2.1,  'Maps session start hour', False, '#666666'),
        (2.35, 'to the TOU rate band.', False, '#666666'),
        (2.75, 'TRUE = approx match for', False, '#666666'),
        (3.0,  'tiered range lookup.', False, '#666666'),
    ]:
        txt(ax, cx + 0.18, cy + ch - 0.35 - dy, line,
            size=7, color=color, bold=bold, zorder=11)

    # Arrow from callout to Rate column header
    ax.annotate('', xy=(9.35, header_y + 0.07),
                xytext=(cx + 0.2, cy + ch - 0.35 - 0.75),
                arrowprops=dict(arrowstyle='->', color='#F4A261', lw=1.5), zorder=12)

    # Legend
    leg_y = 0.35
    rect(ax, 0.1,  leg_y, 0.45, 0.25, '#C8E6C9', radius=0.04)
    txt(ax, 0.65,  leg_y + 0.06, 'Above target',  size=7.5, color='#1A4731')
    rect(ax, 2.2,  leg_y, 0.45, 0.25, '#FFCDD2', radius=0.04)
    txt(ax, 2.75,  leg_y + 0.06, 'Below target',  size=7.5, color='#CC0000')
    rect(ax, 4.3,  leg_y, 0.45, 0.25, '#FFF8E1', radius=0.04, ec='#F4A261', lw=0.8)
    txt(ax, 4.85,  leg_y + 0.06, 'VLOOKUP result', size=7.5, color='#7D4E00')

    f.savefig('assets/images/work/report-rawdata.png', dpi=DPI,
              bbox_inches='tight', facecolor='#FFFFFF')
    plt.close(f)
    print("Saved report-rawdata.png")


build_dashboard()
build_rawdata()
print("All done.")
